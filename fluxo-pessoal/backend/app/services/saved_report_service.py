from __future__ import annotations

from io import BytesIO

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select, update
from sqlalchemy.orm import Session, selectinload

from app.models.report_indicator import ReportIndicator
from app.models.saved_report import SavedReport, SavedReportIndicator
from app.schemas.report_indicator_schema import ReportIndicatorEvaluation
from app.schemas.saved_report_schema import SavedReportCreate, SavedReportEvaluation, SavedReportRead, SavedReportUpdate
from app.services.report_indicator_service import ReportIndicatorService


class SavedReportService:
    def __init__(self, db: Session):
        self.db = db
        self.indicator_service = ReportIndicatorService(db)

    def list(self, include_inactive: bool = False) -> list[SavedReportRead]:
        stmt = (
            select(SavedReport)
            .options(selectinload(SavedReport.indicators).selectinload(SavedReportIndicator.indicator))
            .order_by(SavedReport.display_order, SavedReport.name)
        )
        if not include_inactive:
            stmt = stmt.where(SavedReport.is_active.is_(True))
        return [self._to_read(item) for item in self.db.scalars(stmt)]

    def get(self, report_id: int) -> SavedReport:
        report = self.db.scalar(
            select(SavedReport)
            .options(selectinload(SavedReport.indicators).selectinload(SavedReportIndicator.indicator))
            .where(SavedReport.id == report_id)
        )
        if not report:
            raise HTTPException(status_code=404, detail="Relatório não encontrado")
        return report

    def default(self) -> SavedReport | None:
        return self.db.scalar(
            select(SavedReport)
            .options(selectinload(SavedReport.indicators).selectinload(SavedReportIndicator.indicator))
            .where(SavedReport.is_active.is_(True), SavedReport.is_default_dashboard.is_(True))
            .order_by(SavedReport.display_order)
        )

    def create(self, payload: SavedReportCreate) -> SavedReportRead:
        report = SavedReport(**payload.model_dump(exclude={"indicators"}))
        self.db.add(report)
        self.db.flush()
        if report.is_default_dashboard:
            self._clear_other_defaults(report.id)
        self._replace_indicators(report, payload.indicators)
        self.db.commit()
        return self._to_read(self.get(report.id))

    def update(self, report_id: int, payload: SavedReportUpdate) -> SavedReportRead:
        report = self.get(report_id)
        data = payload.model_dump(exclude_unset=True, exclude={"indicators"})
        for field, value in data.items():
            setattr(report, field, value)
        if report.is_default_dashboard:
            self._clear_other_defaults(report.id)
        if payload.indicators is not None:
            self._replace_indicators(report, payload.indicators)
        self.db.commit()
        return self._to_read(self.get(report.id))

    def delete(self, report_id: int) -> SavedReportRead:
        report = self.get(report_id)
        deleted = self._to_read(report)
        self.db.delete(report)
        self.db.commit()
        return deleted

    def evaluate(self, report_id: int, month: str) -> SavedReportEvaluation:
        report = self.get(report_id)
        indicators = [self.indicator_service._evaluate_indicator(item.indicator, month) for item in report.indicators if item.indicator and item.indicator.is_active]
        return SavedReportEvaluation(
            id=report.id,
            name=report.name,
            description=report.description,
            is_default_dashboard=report.is_default_dashboard,
            indicators=indicators,
        )

    def seed_defaults(self) -> list[SavedReportRead]:
        self.indicator_service.seed_defaults()

        existing = set(self.db.scalars(select(SavedReport.name)))
        indicators = list(self.db.scalars(select(ReportIndicator).where(ReportIndicator.is_active.is_(True)).order_by(ReportIndicator.display_order)))
        created: list[SavedReportRead] = []
        presets = [
            ("Relatório financeiro padrão", "Resumo principal com sobra/falta, custo de vida e capacidade de guardar.", True, [1, 2, 3, 19, 20]),
            ("Relatório de gastos essenciais", "Acompanha essenciais, moradia, saúde, transporte e dívidas.", False, [4, 5, 7, 13, 14, 15]),
            ("Relatório de consumo variável", "Foco em vida flexível, lazer, compras e pressão para ficar negativo.", False, [6, 16, 21, 22, 23]),
            ("Relatório de reservas e objetivos", "Acompanha valores enviados e retirados de reservas.", False, [3, 9, 10]),
            ("Relatório de fórmulas avançadas", "Cenários, percentuais, projeções prováveis e pressão de ficar negativo.", False, [19, 20, 21, 22, 23]),
            ("Relatorio de saude do fluxo de caixa", "Indicadores de comprometimento, pressao, poupanca, renda extra e alerta de caixa negativo.", False, [1, 2, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33]),
        ]
        by_order = {indicator.display_order: indicator for indicator in indicators}
        for order, (name, description, is_default, indicator_orders) in enumerate(presets, start=1):
            if name in existing:
                continue
            payload = SavedReportCreate(
                name=name,
                description=description,
                is_default_dashboard=is_default,
                display_order=order,
                indicators=[
                    {"indicator_id": by_order[indicator_order].id, "position": position}
                    for position, indicator_order in enumerate(indicator_orders)
                    if indicator_order in by_order
                ],
            )
            created.append(self.create(payload))
        return created

    def export_excel(self, report_id: int, month: str) -> BytesIO:
        evaluation = self.evaluate(report_id, month)
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        ws.append(["Relatório", evaluation.name])
        ws.append(["Mês", month])
        ws.append([])
        ws.append(["Indicador", "Rótulo", "Resultado", "Termo", "Valor base", "Probabilidade", "Peso", "Contribuição"])
        for indicator in evaluation.indicators:
            if not indicator.terms:
                ws.append([indicator.name, indicator.result_label, indicator.result, "", "", "", "", ""])
            for term in indicator.terms:
                ws.append([indicator.name, indicator.result_label, indicator.result, term.label, term.amount, term.probability, term.weight, term.contribution])
        self._format_sheet(ws)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _replace_indicators(self, report: SavedReport, indicators) -> None:
        report.indicators.clear()
        for position, indicator in enumerate(indicators):
            report.indicators.append(SavedReportIndicator(indicator_id=indicator.indicator_id, position=indicator.position if indicator.position is not None else position))

    def _clear_other_defaults(self, report_id: int) -> None:
        self.db.execute(update(SavedReport).where(SavedReport.id != report_id).values(is_default_dashboard=False))

    def _to_read(self, report: SavedReport) -> SavedReportRead:
        return SavedReportRead(
            id=report.id,
            name=report.name,
            description=report.description,
            is_default_dashboard=report.is_default_dashboard,
            is_active=report.is_active,
            display_order=report.display_order,
            created_at=report.created_at,
            updated_at=report.updated_at,
            indicators=[self.indicator_service._to_read(item.indicator) for item in report.indicators if item.indicator],
        )

    def _format_sheet(self, ws) -> None:
        header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=5):
            for cell in row:
                if cell.column in {3, 5, 8}:
                    cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
                if cell.column in {6, 7}:
                    cell.number_format = "0.00%"
        for column_cells in ws.columns:
            max_length = max(len("" if cell.value is None else str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 48)
