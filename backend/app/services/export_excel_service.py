from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.account import Account
from app.models.balance_snapshot import BalanceSnapshot
from app.models.chart_account import ChartAccount
from app.models.classification_rule import ClassificationRule
from app.models.enums import ClassificationStatus, TransactionType
from app.models.transaction import Transaction
from app.services.balance_service import BalanceService
from app.services.report_indicator_service import ReportIndicatorService
from app.services.report_service import ReportService
from app.utils.dates import month_bounds


class ExportExcelService:
    def __init__(self, db: Session):
        self.db = db
        self.report_service = ReportService(db)
        self.report_indicator_service = ReportIndicatorService(db)
        self.balance_service = BalanceService(db)

    def export_month(self, month: str) -> BytesIO:
        wb = Workbook()
        wb.remove(wb.active)

        self._summary_sheet(wb, month)
        self._custom_indicators_sheet(wb, month)
        self._accounts_sheet(wb)
        self._transactions_sheet(wb, "Movimentações", month)
        self._chart_accounts_sheet(wb)
        self._categories_sheet(wb, month)
        self._pending_sheet(wb, month)
        self._rules_sheet(wb)
        self._comparison_sheet(wb, month)
        self._typed_transactions_sheet(wb, "Entradas", month, TransactionType.income)
        self._typed_transactions_sheet(wb, "Saídas", month, TransactionType.expense)
        self._typed_transactions_sheet(wb, "Transferências", month, TransactionType.transfer)
        self._typed_transactions_sheet(wb, "Ajustes", month, TransactionType.adjustment)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _summary_sheet(self, wb: Workbook, month: str) -> None:
        ws = wb.create_sheet("Resumo Geral")
        report = self.report_service.monthly(month)
        rows = [
            ("Mês", report.month),
            ("Saldo inicial consolidado", report.opening_balance),
            ("Receita", report.income),
            ("Despesas fixas", report.fixed_expenses),
            ("Despesas variáveis", report.variable_expenses),
            ("Dívidas e obrigações", report.obligations),
            ("Outras despesas", report.other_expenses),
            ("Total de gastos reais", report.total_real_expenses),
            ("Sobra/Falta do mês", report.cash_result),
            ("Transferências internas", report.transfers),
            ("Reservas", report.reserves),
            ("Ajustes", report.adjustments),
            ("Pendências", report.pending_count),
            ("Saldo atual consolidado", report.consolidated_balance),
        ]
        ws.append(["Indicador", "Valor"])
        for row in rows:
            ws.append(list(row))
        self._format_sheet(ws, money_columns={2})

    def _accounts_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Saldos por Conta")
        ws.append(["ID", "Conta", "Instituição", "Tipo", "Saldo inicial", "Saldo calculado", "Saldo salvo", "Ativa"])
        for item in self.balance_service.calculate_all_balances():
            ws.append(
                [
                    item.id,
                    item.name,
                    item.institution,
                    item.account_type.value,
                    item.initial_balance,
                    item.calculated_balance,
                    item.current_balance,
                    "Sim" if item.is_active else "Não",
                ]
            )
        self._format_sheet(ws, money_columns={5, 6, 7})

    def _custom_indicators_sheet(self, wb: Workbook, month: str) -> None:
        ws = wb.create_sheet("Indicadores")
        ws.append(["Indicador", "Rótulo", "Resultado", "Termo", "Operação", "Modo", "Valor base", "Contribuição"])
        for indicator in self.report_indicator_service.evaluate(month, surface="reports"):
            if not indicator.terms:
                ws.append([indicator.name, indicator.result_label, indicator.result, "", "", "", "", ""])
            for term in indicator.terms:
                ws.append(
                    [
                        indicator.name,
                        indicator.result_label,
                        indicator.result,
                        term.label,
                        term.operation.value,
                        term.value_mode.value,
                        term.amount,
                        term.contribution,
                    ]
                )
        self._format_sheet(ws, money_columns={3, 7, 8})

    def _transactions_query(self, month: str):
        start, end = month_bounds(month)
        return (
            select(Transaction, Account.name, ChartAccount.code, ChartAccount.name)
            .join(Account, Account.id == Transaction.account_id)
            .join(ChartAccount, ChartAccount.id == Transaction.chart_account_id, isouter=True)
            .where(Transaction.transaction_date >= start, Transaction.transaction_date < end)
            .order_by(Transaction.transaction_date, Transaction.id)
        )

    def _transactions_sheet(self, wb: Workbook, title: str, month: str) -> None:
        ws = wb.create_sheet(title)
        self._write_transaction_headers(ws)
        for tx, account_name, chart_code, chart_name in self.db.execute(self._transactions_query(month)):
            self._append_transaction(ws, tx, account_name, chart_code, chart_name)
        self._format_sheet(ws, money_columns={6}, date_columns={2})

    def _typed_transactions_sheet(self, wb: Workbook, title: str, month: str, transaction_type: TransactionType) -> None:
        ws = wb.create_sheet(title)
        self._write_transaction_headers(ws)
        stmt = self._transactions_query(month).where(Transaction.transaction_type == transaction_type)
        for tx, account_name, chart_code, chart_name in self.db.execute(stmt):
            self._append_transaction(ws, tx, account_name, chart_code, chart_name)
        self._format_sheet(ws, money_columns={6}, date_columns={2})

    def _pending_sheet(self, wb: Workbook, month: str) -> None:
        ws = wb.create_sheet("Pendentes de Classificação")
        self._write_transaction_headers(ws)
        stmt = self._transactions_query(month).where(Transaction.classification_status == ClassificationStatus.pending)
        for tx, account_name, chart_code, chart_name in self.db.execute(stmt):
            self._append_transaction(ws, tx, account_name, chart_code, chart_name)
        self._format_sheet(ws, money_columns={6}, date_columns={2})

    def _chart_accounts_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Plano de Contas")
        ws.append(["ID", "Código", "Nome", "Conta pai", "Natureza", "Ativa"])
        stmt = select(ChartAccount).order_by(ChartAccount.code)
        for item in self.db.scalars(stmt):
            ws.append([item.id, item.code, item.name, item.parent_id, item.account_nature.value, "Sim" if item.is_active else "Não"])
        self._format_sheet(ws)

    def _categories_sheet(self, wb: Workbook, month: str) -> None:
        ws = wb.create_sheet("Categorias")
        ws.append(["ID categoria", "Código", "Categoria", "Natureza", "Total", "Quantidade"])
        for item in self.report_service.categories(month):
            ws.append([item.chart_account_id, item.code, item.name, item.account_nature, item.total, item.count])
        self._format_sheet(ws, money_columns={5})

    def _rules_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Regras Automáticas")
        ws.append(["ID", "Palavra-chave", "Tipo de match", "Conta contábil", "Tipo", "Prioridade", "Ativa"])
        stmt = (
            select(ClassificationRule, ChartAccount.code, ChartAccount.name)
            .join(ChartAccount, ChartAccount.id == ClassificationRule.chart_account_id)
            .order_by(ClassificationRule.priority, ClassificationRule.keyword)
        )
        for rule, code, name in self.db.execute(stmt):
            ws.append([rule.id, rule.keyword, rule.match_type.value, f"{code} - {name}", rule.transaction_type.value, rule.priority, rule.active])
        self._format_sheet(ws)

    def _comparison_sheet(self, wb: Workbook, month: str) -> None:
        ws = wb.create_sheet("Comparativo Mensal")
        ws.append(["Mês", "Receita", "Fixas", "Variáveis", "Obrigações", "Outras", "Gastos reais", "Sobra/Falta", "Pendências"])
        year = month[:4]
        for item in self.report_service.comparison(f"{year}-01", month):
            ws.append(
                [
                    item.month,
                    item.income,
                    item.fixed_expenses,
                    item.variable_expenses,
                    item.obligations,
                    item.other_expenses,
                    item.total_real_expenses,
                    item.cash_result,
                    item.pending_count,
                ]
            )
        self._format_sheet(ws, money_columns={2, 3, 4, 5, 6, 7, 8})

    def _write_transaction_headers(self, ws) -> None:
        ws.append(
            [
                "ID",
                "Data",
                "Conta",
                "Descrição original",
                "Descrição limpa",
                "Valor",
                "Direção",
                "Tipo",
                "Plano de contas",
                "Status",
                "Origem",
                "Transferência interna",
                "Observações",
            ]
        )

    def _append_transaction(self, ws, tx: Transaction, account_name: str, chart_code: str | None, chart_name: str | None) -> None:
        ws.append(
            [
                tx.id,
                tx.transaction_date,
                account_name,
                tx.description_original,
                tx.description_clean,
                tx.amount,
                tx.direction.value,
                tx.transaction_type.value,
                f"{chart_code} - {chart_name}" if chart_code else "Sem categoria",
                tx.classification_status.value,
                tx.source.value,
                "Sim" if tx.is_internal_transfer else "Não",
                tx.notes,
            ]
        )

    def _format_sheet(self, ws, money_columns: set[int] | None = None, date_columns: set[int] | None = None) -> None:
        money_columns = money_columns or set()
        date_columns = date_columns or set()
        header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column in money_columns:
                    cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
                if cell.column in date_columns:
                    cell.number_format = "dd/mm/yyyy"
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            ws.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 12), 48)
