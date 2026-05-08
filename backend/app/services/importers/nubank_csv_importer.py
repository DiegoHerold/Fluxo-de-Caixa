import csv
from io import BytesIO, StringIO
from zipfile import is_zipfile

import pandas as pd
from openpyxl import load_workbook

from app.models.enums import TransactionSource
from app.services.importers.base_importer import BaseImporter
from app.utils.dates import parse_date
from app.utils.text_normalizer import repair_mojibake


class NubankCSVImporter(BaseImporter):
    source = TransactionSource.nubank_csv

    def parse(self, content: bytes, filename: str, account_id: int) -> list[dict]:
        df = self._read_statement(content)

        date_col = self.find_column(df, ["date", "data"])
        description_col = self.find_column(df, ["description", "descricao", "descrição", "title", "detalhes"])
        amount_col = self.find_column(df, ["amount", "valor", "value"])
        external_col = self.find_column(df, ["id", "identificador", "external_id"], required=False)

        payloads: list[dict] = []
        for _, row in df.iterrows():
            if pd.isna(row[date_col]) or pd.isna(row[amount_col]) or pd.isna(row[description_col]):
                continue
            amount = self.amount(row[amount_col])
            payloads.append(
                self.build_payload(
                    account_id=account_id,
                    transaction_date=parse_date(row[date_col]),
                    description_original=str(row[description_col]),
                    amount=amount,
                    external_id=str(row[external_col]).strip() if external_col and pd.notna(row[external_col]) else None,
                )
            )
        return payloads

    def _read_statement(self, content: bytes) -> pd.DataFrame:
        if is_zipfile(BytesIO(content)):
            return self._read_excel_wrapped_csv(content)

        last_error: Exception | None = None
        for encoding in ("utf-8", "utf-8-sig", "latin1"):
            try:
                text = repair_mojibake(content.decode(encoding))
                return pd.read_csv(StringIO(text), sep=None, engine="python")
            except Exception as exc:
                last_error = exc
        raise ValueError(f"Não foi possível ler o CSV do Nubank: {last_error}")

    def _read_excel_wrapped_csv(self, content: bytes) -> pd.DataFrame:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        lines: list[str] = []
        for row in worksheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if len(values) == 1:
                line = values[0]
            else:
                line = ",".join(values)
            if line.strip():
                lines.append(repair_mojibake(line))

        reader = csv.reader(lines)
        rows = list(reader)
        if not rows:
            raise ValueError("Planilha CSV do Nubank sem linhas")
        return pd.DataFrame(rows[1:], columns=[repair_mojibake(column) for column in rows[0]])
